# -*- coding: UTF-8 -*-
import  cv2
import  random
import numpy as np
import xlsxwriter
import  math
import skimage
import openpyxl
import matplotlib.pyplot as plt


def sp_noise(image, prob):
    '''
    添加椒盐噪声
    :param image: 灰度图
    :param prob:噪声比例
    '''
    output = np.zeros(image.shape, np.uint8)
    for i in range(image.shape[0]):
        for j in range(image.shape[1]):
            r1 = random.random()
            r2 = random.random()
            if r1 < prob:
                if r2 >= 0.5:
                    output[i][j] = 0
                else:
                    output[i][j] = 255
            else:
                output[i][j] = image[i][j]
    return output


def mindiff(image, i, j):
    '''
    寻找与噪声点的像素值最接近的信号点
    :param image: 灰度图
    :param i: 横坐标
    :param j: 纵坐标
    :return: 像素值
    '''
    x = [
        abs(int(image[i][j]) - int(image[i - 1][j - 1])), abs(int(image[i][j]) - int(image[i - 1][j])),
        abs(int(image[i][j]) - int(image[i - 1][j + 1])), abs(int(image[i][j]) - int(image[i][j - 1])),
        abs(int(image[i][j]) - int(image[i][j + 1])), abs(int(image[i][j]) - int(image[i + 1][j - 1])),
        abs(int(image[i][j]) - int(image[i + 1][j])), abs(int(image[i][j]) - int(image[i + 1][j + 1])),
        ]
    while 0 in x:  #不选用噪声点
        x.remove(0)
    return min(x)


def neighbor_4_mean(image, i, j):
    '''
    :param image: 原图像素值
    :return: 4邻域均值
    '''
    x = [
        image[i - 1][j],
        image[i][j - 1], image[i][j + 1],
        image[i + 1][j],
        ]
    return np.mean(x)


def neighbor_8_mean(image, i, j):
    '''
    :param image: 原图像素值
    :return: 8邻域均值
    '''
    x = [
        image[i - 1][j - 1], image[i - 1][j], image[i - 1][j + 1],
        image[i][j - 1],     image[i][j + 1],
        image[i + 1][j - 1], image[i + 1][j], image[i + 1][j + 1]
        ]
    return np.mean(x)


def median_4(image, i, j):
    '''
    :param image: 原图像素值
    :return: 4邻域中值
    '''
    x = [
        image[i - 1][j],
        image[i][j - 1], image[i][j + 1],
        image[i + 1][j],
        ]
    x.sort()
    return x


def median_8(image, i, j):
    '''
    :param image: 原图像素值
    :return: 8邻域中值
    '''
    x = [
        image[i - 1][j - 1], image[i - 1][j], image[i - 1][j + 1],
        image[i][j - 1],     image[i][j + 1],
        image[i + 1][j - 1], image[i + 1][j], image[i + 1][j + 1]
        ]
    x.sort()
    return x


def cal_psnr(image, denoise_image):
    '''
    计算PSNR
    :param image:原图
    :param denoise_image:去噪图
    :return: PSNR
    '''
    m = image.shape[0]
    n = image.shape[1]
    for i in range(m):
        for j in range(n):
            sum += (image[i][j] - denoise_image[i][j]) * (image[i][j] - denoise_image[i][j])
    MSE = sum / (m * n)
    psnr = 10 * math.log(255 * 255 / MSE, 10)
    return psnr


def neighbor4_3coef(image, i, j, list):
    return list[0] * image[i][j-1] + list[1] * image[i-1][j] + list[2] * image[i][j+1] + (1 - list[0] - list[1] - list[2]) * image[i+1][j]


def neighbor4_4coef(image, i, j, list):
    return list[0] * image[i][j-1] + list[1] * image[i-1][j] + list[2] * image[i][j+1] + list[3] * image[i+1][j]


def neighbor4_3coef_denoise(image, i, j, list):
    x = [
        image[i - 1][j],
        image[i][j - 1], image[i][j + 1],
        image[i + 1][j],
        ]
    y = []
    sum = 0
    for i in range(len(x)):
        if x[i] == 0 or x[i] == 255:
            y.append(0)
        else:
            y.append(1)
            sum += list[i]
    if np.sum(y) == 0:
        return image[i][j]
    z = []
    for i in range(len(y)):
        if y[i] == 0:
            z.append(0)
        else:
            z.append(list[i]/sum)
    return z[0] * image[i][j - 1] + z[1] * image[i - 1][j] + z[2] * image[i][j + 1] + z[3] * image[i + 1][j]

def has_edge_point(edge, i, j):
    #if edge.cell(i+1,j+1).value[-1] == '5':
    for k in range(-1,2):
        for l in range(-1,2):
            if edge.cell(i+k,j+l).value[-1] == '5':
                return 1
            else:
                continue
    return 0


img = cv2.imread('C:/Users/AMDyes/Desktop/result/lena512.bmp', cv2.IMREAD_GRAYSCALE) #读取灰度图
filename = 'C:/Users/AMDyes/Desktop/result/edge.xlsx'
wb = openpyxl.load_workbook(filename)
edge = wb['Sheet1']
workbook0 = xlsxwriter.Workbook('C:/Users/AMDyes/Desktop/result/lena512_original.xlsx') # 原图像素值
worksheet0 = workbook0.add_worksheet()
for i in range(img.shape[0]):
    for j in range(img.shape[1]):
        worksheet0.write(i, j, str(img[i][j]))
workbook0.close()

# print("img_cv:",img.shape)
# cv2.imshow('original_img',img)
# cv2.waitKey(0)

#直方图
'''

plt.figure("lena")
arr = img.flatten()  # 一维数组
n, bins, patches = plt.hist(arr, bins = 256, density=1, facecolor='green', alpha = 0.75)
plt.show()

'''

noise_density = 0.1
out1 = sp_noise(img, noise_density)
cv2.imwrite('C:/Users/AMDyes/Desktop/result/noise_img10%.bmp', out1)


#初步处理噪声
#out2 = np.pad(out1, ((1,1),(1,1)),mode = 'constant')  # 填充边界
#out3 = out2

# 3-coefficients
list1 = [0.09196457, 0.40794827, 0.09194055, 0.40814661] #包含边缘点
list2 = [0.12258222, 0.3728235, 0.12610541, 0.37848887] #不含边缘点

# 4-coefficients
list3 = [0.09260788, 0.40785214, 0.09245374, 0.40815818]  #包含边缘点
list4 = [0.12266286, 0.37281655, 0.12610221, 0.37853621]  #不含边缘点

out2 = out1
for i in range(1, out2.shape[0] - 1):
    for j in range(1,out2.shape[1] - 1):
        if(out2[i][j] == 0):
            out1[i][j] = neighbor4_3coef_denoise(out2, i, j, list1)
        elif(out2[i][j] == 255):
            out1[i][j] = neighbor4_3coef_denoise(out2, i, j, list1)

for k in range(10):
    out2 = out1
    for i in range(1, out2.shape[0] - 1):
        for j in range(1, out2.shape[1] - 1):
            if (out2[i][j] == 0):
                out1[i][j] = neighbor4_3coef_denoise(out2, i, j, list1)
            elif (out2[i][j] == 255):
                out1[i][j] = neighbor4_3coef_denoise(out2, i, j, list1)
cv2.imshow('denoise',out1)
cv2.waitKey(0)


out3 = out4 = img.astype(np.int16)
for i in range(1, out3.shape[0] - 1):
    for j in range(1,out3.shape[1] - 1):
        x = median_8(out4, i, j)
        out3[i][j] = x[3] / 2 + x[4] / 2




'''
# out3为处理好的去噪图
workbook2 = xlsxwriter.Workbook('C:/Users/AMDyes/Desktop/result/lena512_' + str(int(noise_density * 200)) + '%_diff.xlsx') # 建立文件
worksheet2 = workbook2.add_worksheet()
workbook3 = xlsxwriter.Workbook('C:/Users/AMDyes/Desktop/result/lena512_' + str(int(noise_density * 200)) + '%_process.xlsx') # 建立文件
worksheet3 = workbook3.add_worksheet()
cell_format1 = workbook3.add_format()
cell_format1.set_pattern(1)
cell_format1.set_bg_color('red')

cell_format2 = workbook3.add_format()
cell_format2.set_pattern(1)
cell_format2.set_bg_color('green')

# 去除填充边界,out4为处理好的去噪图
out4 = np.zeros(img.shape, np.uint8)
for i in range(out4.shape[0]):
    for j in range(out4.shape[1]):
        out4[i][j] = out3[i+1][j+1]
# 原图与去噪图差异
diff = cv2.absdiff(img, out4)
#print("img_cv:",diff.shape)
for i in range(diff.shape[0]):
    for j in range(diff.shape[0]):
        worksheet2.write(i, j, str(diff[i][j]))
workbook2.close()
for i in range(out4.shape[0]):
    for j in range(out4.shape[1]):
        if(abs(diff[i][j] >= 10)):
            worksheet3.write(i, j, str(out4[i][j]), cell_format1)
        elif(abs(diff[i][j] >= 5)):
            worksheet3.write(i, j, str(out4[i][j]), cell_format2)
        else:
            worksheet3.write(i, j, str(out4[i][j]))
workbook3.close()
'''





# 计算psnr和ssim
psnr = skimage.measure.compare_psnr(img, out1, data_range = 255)
ssim = skimage.measure.compare_ssim(img, out1, data_range = 255)

print(ssim,psnr)










